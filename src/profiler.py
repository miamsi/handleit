import openpyxl
from typing import Dict, Any

class ExcelProfiler:
    @staticmethod
    def profile_workbook(file_path: str) -> Dict[str, Any]:
        """Analyzes Excel topology without loading everything into memory."""
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        profile = {"sheets": {}}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info = {
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "potential_header_row": None,
                "sample_columns": []
            }
            
            # Scan first 20 rows to find the header (row with most string values)
            max_strings = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row if ws.max_row else 20), values_only=True), 1):
                string_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
                if string_count > max_strings:
                    max_strings = string_count
                    sheet_info["potential_header_row"] = row_idx
                    sheet_info["sample_columns"] = [str(cell)[:50] if cell else "" for cell in row]
            
            profile["sheets"][sheet_name] = sheet_info
            
        wb.close()
        return profile
